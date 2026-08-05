# -*- coding: utf-8 -*-
"""
流水预处理 + 延迟清算检测模块

功能：
  1. preprocess_transactions — 将多个账户混合在同一个 sheet 的流水，
     按账户拆分为独立 sheet，每个 sheet 内的交易按日期+时间排序，保存为副本。
  2. detect_end_of_day_nonzero — 延迟清算检测：对所有账户、所有日期，
     筛选出当日最后一笔交易余额不为零的记录。
"""

import os
import re
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


# ── 列名别名映射（从 column_mapping.json 加载，与 process_zero_balance.py 共享） ──
# 修改列名对应关系请编辑 column_mapping.json 文件，无需改动代码。

import json
_PREPROCESS_CONFIG_CACHE = None

def _get_column_groups():
    """从 column_mapping.json 加载预处理列名分组，失败时使用硬编码默认值。"""
    global _PREPROCESS_CONFIG_CACHE
    if _PREPROCESS_CONFIG_CACHE is not None:
        return _PREPROCESS_CONFIG_CACHE

    defaults = {
        'account': ['账户', '账号', '帐号'],
        'date': ['交易日期', '入帐日期', '入账日期', '日期'],
        'time': ['交易时间', '时间', '交易时刻'],
        'balance': ['余额', '账户余额'],
        'amount': ['交易金额', '发生额', '发生额（分）', '金额'],
    }

    config_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'column_mapping.json')
    if os.path.exists(config_path):
        try:
            with open(config_path, 'r', encoding='utf-8') as f:
                loaded = json.load(f)
            groups = loaded.get("preprocess_column_groups", defaults)
            # 过滤掉以 _ 开头的注释键（如 __comment__）
            _PREPROCESS_CONFIG_CACHE = {k: v for k, v in groups.items() if not k.startswith('_')}
            return _PREPROCESS_CONFIG_CACHE
        except Exception:
            pass
    _PREPROCESS_CONFIG_CACHE = defaults
    return _PREPROCESS_CONFIG_CACHE


def _find_column(df, aliases):
    """在 DataFrame 列名中查找第一个匹配别名的列，返回列名或 None。"""
    for col in df.columns:
        col_clean = str(col).strip().lstrip('\ufeff')
        if col_clean in aliases:
            return col
    return None


def _detect_header_row(xls, sheet_name, max_rows=15):
    """自动检测表头所在行（与 process_zero_balance.py 相同逻辑）。"""
    best_row = 0
    best_score = -1.0

    for h in range(max_rows):
        try:
            df = pd.read_excel(xls, sheet_name=sheet_name, header=h, nrows=0)
            cols = [str(c).strip() for c in df.columns if pd.notna(c)]
            if not cols:
                continue

            unique_ratio = len(set(cols)) / len(cols) if cols else 0

            score = 0
            for c in cols:
                if '入帐' in c or ('日期' in c and '交易' in c):
                    score += 3
                elif '日期' in c or '借贷标志' in c:
                    score += 3
                elif '发生额' in c or ('金额' in c and '交易' in c):
                    score += 3
                elif '金额' in c:
                    score += 2
                elif '户名' in c or '行名' in c:
                    score += 2
                elif '余额' in c:
                    score += 2
                elif '摘要' in c or '描述' in c:
                    score += 1
                elif '账号' in c or '帐号' in c:
                    score += 1
                elif '对方' in c:
                    score += 1
                elif '记录号' in c:
                    score += 1
            score = score * unique_ratio

            if score > best_score:
                best_score = score
                best_row = h
        except Exception:
            continue

    return best_row


def _safe_str(v):
    """将值安全转换为字符串。"""
    if pd.isna(v):
        return ''
    if isinstance(v, (int, float)):
        # 防止科学计数法（如 1.82E+16）
        return str(int(float(v)))
    return str(v).strip()


def _parse_datetime(date_val, time_val=None):
    """
    尝试解析日期和时间为 datetime 对象。
    返回 (datetime_obj, 是否成功)
    """
    dt = None
    # 先解析日期
    if pd.isna(date_val):
        return None, False

    if isinstance(date_val, datetime):
        dt = date_val
    elif isinstance(date_val, (int, float)):
        try:
            s = str(int(date_val))
            dt = datetime.strptime(s, '%Y%m%d')
        except (ValueError, TypeError):
            pass
    elif isinstance(date_val, str):
        s_raw = date_val.strip()
        s = s_raw.replace('-', '').replace('/', '').replace(' ', '').replace(':', '')
        if s.isdigit() and len(s) >= 8:
            try:
                dt = datetime.strptime(s[:8], '%Y%m%d')
            except ValueError:
                pass
        if dt is None:
            for sep in ('/', '-'):
                if sep in s_raw:
                    parts = s_raw.split(sep)
                    if len(parts) == 3:
                        try:
                            dt = datetime(int(parts[0]), int(parts[1]), int(parts[2]))
                        except (ValueError, TypeError):
                            pass

    if dt is None:
        return None, False

    # 如果有时分秒，拼上去
    if time_val is not None and not pd.isna(time_val):
        t_str = str(time_val).strip().replace('：', ':')
        # 尝试多种时间格式
        t_patterns = ['%H:%M:%S', '%H:%M', '%H%M%S', '%H%M']
        for fmt in t_patterns:
            try:
                t_obj = datetime.strptime(t_str[:len(fmt)], fmt)
                dt = dt.replace(hour=t_obj.hour, minute=t_obj.minute, second=t_obj.second)
                break
            except (ValueError, IndexError):
                continue

    return dt, True


def _is_excel_file(filepath):
    ext = os.path.splitext(filepath)[1].lower()
    return ext in ('.xlsx', '.xls')


def preprocess_transactions(input_path, output_path=None):
    """
    流水文件预处理：将多账户混合的 Excel 按账户拆分、排序，保存为副本。

    处理逻辑：
    1. 自动检测表头行，读取数据
    2. 自动识别"账户"、"交易日期"、"交易时间"、"余额"等关键列
    3. 按账户分组，每个账户的数据按（交易日期, 交易时间）升序排列
    4. 每个账户写入一个独立 Sheet，Sheet 名为账户名（截取前 31 字符）

    参数
    ----------
    input_path : str
        输入的 Excel 文件路径。
    output_path : str, optional
        输出文件路径。默认为输入文件名 + "_预处理.xlsx"。

    返回
    -------
    str
        输出文件路径。若处理失败返回 None。
    """
    if not os.path.exists(input_path):
        raise FileNotFoundError(f"输入文件不存在: {input_path}")

    if not _is_excel_file(input_path):
        raise ValueError(f"仅支持 Excel 文件（.xlsx/.xls）：{input_path}")

    if output_path is None:
        base, ext = os.path.splitext(input_path)
        output_path = f"{base}_预处理{ext}"

    # 读取 Excel
    xls = pd.ExcelFile(input_path)
    first_sheet = xls.sheet_names[0]

    # 自动检测表头
    header_row = _detect_header_row(xls, first_sheet)
    df_raw = pd.read_excel(xls, sheet_name=first_sheet, header=header_row)

    # 去除全空行
    df_raw = df_raw.dropna(how='all').reset_index(drop=True)

    if len(df_raw) == 0:
        raise ValueError("未读取到任何数据")

    # ── 自动识别关键列（从 column_mapping.json 加载别名配置） ──
    groups = _get_column_groups()
    acct_col = _find_column(df_raw, set(groups['account']))
    date_col = _find_column(df_raw, set(groups['date']))
    time_col = _find_column(df_raw, set(groups['time']))
    balance_col = _find_column(df_raw, set(groups['balance']))
    amount_col = _find_column(df_raw, set(groups['amount']))

    # 列出实际列名，方便调试
    actual_cols = list(df_raw.columns)
    print(f"  检测到列名: {actual_cols}")
    print(f"  账户列: {acct_col}")
    print(f"  日期列: {date_col}")
    print(f"  时间列: {time_col}")
    print(f"  余额列: {balance_col}")

    if acct_col is None:
        groups = _get_column_groups()
        raise ValueError(f"未找到账户列（尝试识别: {groups['account']}）")
    if date_col is None:
        raise ValueError(f"未找到日期列（尝试识别: {groups['date']}）")

    # ── 处理账户号仅在首行出现的情况（如零余额账户格式） ──
    # 检查账户列是否有大量 NaN，若是则从第一个非空值开始向前填充
    acct_non_null_count = df_raw[acct_col].notna().sum()
    if acct_non_null_count > 0 and acct_non_null_count < len(df_raw) * 0.3:
        # 少于30%的行有账户值 → 前向填充
        df_raw[acct_col] = df_raw[acct_col].ffill()
        print(f"  账户列仅有 {acct_non_null_count} 行有值，已执行前向填充")

    # ── 解析账户、日期、时间 ──
    records = []
    skipped_no_acct = 0
    skipped_no_date = 0

    for idx, row in df_raw.iterrows():
        # 账户
        acct_val = row.get(acct_col)
        acct_str = _safe_str(acct_val)
        if not acct_str:
            skipped_no_acct += 1
            continue

        # 日期
        date_val = row.get(date_col)
        time_val = row.get(time_col) if time_col else None
        dt, ok = _parse_datetime(date_val, time_val)
        if not ok:
            skipped_no_date += 1
            continue

        # 余额
        balance_val = None
        if balance_col:
            try:
                balance_val = pd.to_numeric(row.get(balance_col), errors='coerce')
                if pd.isna(balance_val):
                    balance_val = None
            except Exception:
                pass

        # 组装行记录
        record = {'_账户': acct_str, '_日期时间': dt}
        # 保留所有原始列
        for col in df_raw.columns:
            record[col] = row[col]
        # 额外保留解析好的结构化字段
        if balance_val is not None:
            record['_余额_数值'] = balance_val
        records.append(record)

    if skipped_no_acct:
        print(f"  跳过无账户行: {skipped_no_acct}")
    if skipped_no_date:
        print(f"  跳过无日期行: {skipped_no_date}")
    if not records:
        raise ValueError("没有有效的交易记录")

    df = pd.DataFrame(records)

    # ── 按账户分组，排序 ──
    grouped = df.groupby('_账户', sort=False)
    account_list = list(grouped.groups.keys())
    print(f"  发现 {len(account_list)} 个账户: {account_list}")

    # ── 写出到新 Excel ──
    from openpyxl import Workbook

    wb = Workbook()
    # 删除默认 sheet
    wb.remove(wb.active)

    # 原始列顺序（排除 _ 开头的辅助列）
    output_columns = [c for c in df_raw.columns]

    for acct in account_list:
        group = grouped.get_group(acct).copy()
        # 按日期时间排序
        group = group.sort_values('_日期时间', ascending=True).reset_index(drop=True)

        # Sheet 名取账户名（Excel 限制 31 字符）
        sheet_name = str(acct)[:31]
        # 处理非法字符
        illegal_chars = r'[\\/*?:\[\]]'
        sheet_name = re.sub(illegal_chars, '_', sheet_name)

        ws = wb.create_sheet(title=sheet_name)

        # 写入表头
        for col_idx, col_name in enumerate(output_columns, 1):
            ws.cell(row=1, column=col_idx, value=col_name)

        # 写入数据
        for row_idx, (_, record) in enumerate(group.iterrows(), 2):
            for col_idx, col_name in enumerate(output_columns, 1):
                val = record[col_name]
                if pd.isna(val):
                    ws.cell(row=row_idx, column=col_idx, value=None)
                else:
                    ws.cell(row=row_idx, column=col_idx, value=val)

        print(f"  Sheet [{sheet_name}]: {len(group)} 条记录")

    wb.save(output_path)
    print(f"\n  预处理完成，已保存: {output_path}")
    return output_path


def _read_flow_records(filepath):
    """读取单个流水文件中的全部记录（Excel 多 Sheet 或 CSV/TXT）。"""
    groups = _get_column_groups()

    if _is_excel_file(filepath):
        xls = pd.ExcelFile(filepath)
        all_records = []
        for sheet_name in xls.sheet_names:
            header_row = _detect_header_row(xls, sheet_name)
            df = pd.read_excel(xls, sheet_name=sheet_name, header=header_row)
            df = df.dropna(how='all').reset_index(drop=True)
            if len(df) > 0:
                all_records.extend(_extract_account_records(df, groups, sheet_name))
        return all_records

    df = None
    for encoding in ('utf-8', 'gbk', 'gb2312', 'utf-16'):
        try:
            df = pd.read_csv(filepath, encoding=encoding)
            break
        except (UnicodeDecodeError, UnicodeError):
            continue
    if df is None:
        raise ValueError(f"无法识别文件编码: {filepath}")
    df = df.dropna(how='all').reset_index(drop=True)
    return _extract_account_records(df, groups, os.path.basename(filepath))


def _extract_account_records(df, groups, source_label):
    """从单个 DataFrame 中按账号/日期/时间提取流水记录。"""
    acct_col = _find_column(df, set(groups['account']))
    date_col = _find_column(df, set(groups['date']))
    time_col = _find_column(df, set(groups['time']))

    if acct_col is None:
        raise ValueError(f"文件 {source_label} 未找到账户列（尝试识别: {groups['account']}）")
    if date_col is None:
        raise ValueError(f"文件 {source_label} 未找到日期列（尝试识别: {groups['date']}）")

    acct_non_null_count = df[acct_col].notna().sum()
    if acct_non_null_count > 0 and acct_non_null_count < len(df) * 0.3:
        df[acct_col] = df[acct_col].ffill()

    records = []
    for idx, row in df.iterrows():
        acct_str = _safe_str(row.get(acct_col))
        if not acct_str:
            continue
        time_val = row.get(time_col) if time_col else None
        dt, ok = _parse_datetime(row.get(date_col), time_val)
        if not ok:
            continue
        record = {'_账户': acct_str, '_日期时间': dt}
        for col in df.columns:
            record[col] = row[col]
        records.append(record)
    return records


def _safe_filename(name):
    """将账号转为安全的文件名。"""
    name = str(name).strip()
    name = re.sub(r'[\\/:*?"<>|\r\n\t]+', '_', name)
    name = name.strip('. ')
    return name or '未知账号'


def preprocess_transactions_folder(input_folder, output_dir=None):
    """
    文件夹流水按账号拆分：读取文件夹内所有流水文件，按账号分组，
    每个账号按原日期时间排序后保存为独立 Excel 文件，文件名为账号。
    """
    if not os.path.isdir(input_folder):
        raise NotADirectoryError(f"流水文件夹不存在: {input_folder}")

    if output_dir is None:
        output_dir = os.path.join(input_folder, '按账号拆分')
    os.makedirs(output_dir, exist_ok=True)

    exts = ('.csv', '.txt', '.xlsx', '.xls')
    files = sorted(
        os.path.join(input_folder, name)
        for name in os.listdir(input_folder)
        if os.path.splitext(name)[1].lower() in exts
    )
    if not files:
        raise ValueError(f"文件夹中没有流水文件: {input_folder}")

    all_records = []
    for filepath in files:
        print(f"  读取文件: {filepath}")
        all_records.extend(_read_flow_records(filepath))

    if not all_records:
        raise ValueError("未读取到任何有效的流水记录")

    df = pd.DataFrame(all_records)
    grouped = df.groupby('_账户', sort=False)
    output_columns = [c for c in df.columns if not c.startswith('_')]
    output_paths = []

    for acct in grouped.groups:
        group = grouped.get_group(acct).copy()
        group = group.sort_values('_日期时间', kind='stable', ascending=True).reset_index(drop=True)
        filename = f"{_safe_filename(acct)}.xlsx"
        output_path = os.path.join(output_dir, filename)
        group[output_columns].to_excel(output_path, index=False)
        output_paths.append(output_path)
        print(f"  账号 [{acct}] -> {output_path} ({len(group)} 条)")

    print(f"\n  文件夹拆分完成，共 {len(output_paths)} 个账号文件")
    print(f"  输出目录: {output_dir}")
    return output_paths


def detect_finance_account_verification(input_path, keywords=None):
    """
    财政专户校验：筛选对方户名包含指定关键字的交易，按对方户名汇总数量。
    """
    if keywords is None:
        keywords = ['待报解', '待结算', '非税收入', '暂收款', '暂付款']

    if os.path.isdir(input_path):
        exts = ('.csv', '.txt', '.xlsx', '.xls')
        files = sorted(
            os.path.join(input_path, name)
            for name in os.listdir(input_path)
            if os.path.splitext(name)[1].lower() in exts
        )
        if not files:
            raise ValueError(f"文件夹中没有流水文件: {input_path}")
    else:
        files = [input_path]

    all_records = []
    for filepath in files:
        print(f"  读取文件: {filepath}")
        all_records.extend(_read_flow_records(filepath))

    if not all_records:
        raise ValueError("未读取到任何有效的流水记录")

    df = pd.DataFrame(all_records)
    cpty_col = _find_column(df, {'对方户名', '对方行名', '付款人名称', '收款人名称'})
    if cpty_col is None:
        raise ValueError("未找到对方户名/付款人名称/收款人名称列")

    groups = _get_column_groups()
    amt_col = _find_column(df, set(groups['amount']))
    pattern = '|'.join(re.escape(keyword) for keyword in keywords)
    mask = df[cpty_col].astype(str).str.contains(pattern, na=False, regex=True)
    filtered = df.loc[mask]
    if len(filtered) == 0:
        columns = ['对方户名', '交易笔数']
        if amt_col:
            columns.append('金额合计')
        return pd.DataFrame(columns=columns)

    result = filtered.groupby(cpty_col, sort=False).agg(
        交易笔数=(cpty_col, 'size'),
    ).reset_index()
    result = result.rename(columns={cpty_col: '对方户名'})
    if amt_col:
        result['金额合计'] = filtered.groupby(cpty_col)[amt_col].apply(
            lambda s: pd.to_numeric(s, errors='coerce').sum()
        ).reset_index(drop=True)
        result = result[['对方户名', '交易笔数', '金额合计']]
    result = result.sort_values('交易笔数', ascending=False).reset_index(drop=True)
    return result


def preprocess_transactions_floder(input_folder, output_dir=None):
    """兼容历史拼写的别名，等价于 preprocess_transactions_folder。"""
    return preprocess_transactions_folder(input_folder, output_dir)


def detect_end_of_day_nonzero(input_path, output_path=None):
    """
    延迟清算检测：对所有账户、所有日期，筛选出当日最后一笔交易余额不为零的记录。

    处理逻辑：
    1. 读取预处理后的 Excel（每个账户一个 Sheet，已按时间排序）
    2. 对每个账户 Sheet：
       a. 查找"交易日期"和"交易时间"列（用于分组和排序）
       b. 按交易日期分组
       c. 每个日期内按交易时间排序，取最后一笔
       d. 检查该笔余额是否不为 0
       e. 若不为 0，记录该行
    3. 汇总输出到结果 Excel

    参数
    ----------
    input_path : str
        预处理后的 Excel 文件路径（每个账户一个 Sheet）。
    output_path : str, optional
        输出文件路径。默认为输入文件所在目录 + "延迟清算结果.xlsx"。

    返回
    -------
    pd.DataFrame
        包含所有符合条件的记录。
    """
    if not os.path.exists(input_path):
        raise FileNotFoundError(f"输入文件不存在: {input_path}")

    if output_path is None:
        base_dir = os.path.dirname(input_path) or '.'
        output_path = os.path.join(base_dir, "延迟清算结果.xlsx")

    xls = pd.ExcelFile(input_path)
    sheet_names = xls.sheet_names

    all_results = []

    for sheet_name in sheet_names:
        header_row = _detect_header_row(xls, sheet_name)
        df = pd.read_excel(xls, sheet_name=sheet_name, header=header_row)
        df = df.dropna(how='all').reset_index(drop=True)

        if len(df) == 0:
            continue

        # 识别列（从 column_mapping.json 加载别名配置）
        groups = _get_column_groups()
        date_col = _find_column(df, set(groups['date']))
        time_col = _find_column(df, set(groups['time']))
        balance_col = _find_column(df, set(groups['balance']))
        amt_col = _find_column(df, set(groups['amount']))

        print(f"  Sheet [{sheet_name}]: 日期={date_col}, 时间={time_col}, 余额={balance_col}")

        if date_col is None or balance_col is None:
            print(f"    跳过: 缺少日期列或余额列")
            continue

        # 解析日期和余额
        parsed = []
        for idx, row in df.iterrows():
            date_val = row.get(date_col)
            time_val = row.get(time_col) if time_col else None
            dt, ok = _parse_datetime(date_val, time_val)
            if not ok:
                continue

            try:
                bal = pd.to_numeric(row.get(balance_col), errors='coerce')
                if pd.isna(bal):
                    continue
            except Exception:
                continue

            parsed.append({
                'row_idx': idx,
                'date_obj': dt,
                'balance': bal,
                'original_row': row,
                'sheet_name': sheet_name,
            })

        if not parsed:
            continue

        df_parsed = pd.DataFrame(parsed)

        # 按日分组，取当日最后一笔（按时间排序后取最后一条）
        # 如果有时分秒，按时间排序；否则按文件行序
        date_groups = df_parsed.groupby(df_parsed['date_obj'].dt.date)

        for day, group in date_groups:
            # 如果有时分秒，按时间排序取最后一条
            if time_col:
                group_sorted = group.sort_values('date_obj', ascending=True)
            else:
                group_sorted = group.sort_values('row_idx', ascending=True)

            last_row = group_sorted.iloc[-1]
            bal = last_row['balance']

            if bal is None or abs(bal) < 0.001:
                continue  # 余额为零，跳过

            # 构建结果记录
            orig = last_row['original_row']
            result = {
                '账户': sheet_name,
                '日期': last_row['date_obj'].strftime('%Y-%m-%d'),
                '余额': round(bal, 2),
            }

            # 保留原始列中的有用信息
            if date_col:
                result['原始日期'] = str(orig.get(date_col, ''))
            if time_col:
                result['交易时间'] = str(orig.get(time_col, ''))
            if amt_col:
                try:
                    result['交易金额'] = round(pd.to_numeric(orig.get(amt_col), errors='coerce'), 2)
                except Exception:
                    result['交易金额'] = orig.get(amt_col, '')

            # 也保留其他主要列（摘要、对方户名等）
            for extra_col in ['摘要', '对方户名', '对方行名', '对方账号', '交易描述']:
                if extra_col in df.columns:
                    val = orig.get(extra_col, '')
                    result[extra_col] = str(val) if not pd.isna(val) else ''

            all_results.append(result)

    results_df = pd.DataFrame(all_results)

    # 输出
    if len(results_df) > 0:
        # 排序：按账户、日期
        results_df = results_df.sort_values(['账户', '日期'], ascending=[True, True]).reset_index(drop=True)

    # 保存
    results_df.to_excel(output_path, index=False)
    print(f"\n  延迟清算检测完成，共 {len(results_df)} 条记录")
    print(f"  结果已保存: {output_path}")

    return results_df


# ── 独立运行入口 ──

if __name__ == "__main__":
    import sys

    if len(sys.argv) < 2:
        print("用法:")
        print("  python preprocess.py <Excel文件>            # 预处理")
        print("  python preprocess.py --delayed <Excel文件>  # 延迟清算检测")
        sys.exit(1)

    if sys.argv[1] == '--delayed' and len(sys.argv) >= 3:
        detect_end_of_day_nonzero(sys.argv[2])
    else:
        preprocess_transactions(sys.argv[1])
